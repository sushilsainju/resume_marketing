import io
import csv
import os
from openpyxl import load_workbook

class FileProcessor:
    def __init__(self, file_storage):
        self.file_storage = file_storage

    def parse_recruiters(self):
        filename = getattr(self.file_storage, 'filename', None) or getattr(self.file_storage, 'name', None)
        if not filename:
            raise ValueError("Uploaded file has no filename attribute")
        filename = os.path.basename(filename).lower()
        if filename.endswith('.csv'):
            return self._parse_csv()
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file type")

    def _parse_csv(self):
        # Ensure the pointer is at the start
        if hasattr(self.file_storage, 'seek'):
            self.file_storage.seek(0)
        # Read the file content
        content = self.file_storage.read()
        # If it's bytes, decode to string
        if isinstance(content, bytes):
            content = content.decode('utf-8')
        stream = io.StringIO(content)
        reader = csv.DictReader(stream)
        recruiters = []
        for row in reader:
            recruiters.append({
                'first_name': row.get('first_name') or row.get('First Name') or '',
                'email': row.get('email') or row.get('Email') or '',
                'company_name': row.get('company_name') or row.get('Company') or ''
            })
        return recruiters

    def _parse_excel(self):
        stream = io.BytesIO(self.file_storage.stream.read())
        workbook = load_workbook(filename=stream, read_only=True)
        sheet = workbook.active
        headers = [str(cell).lower() if cell else '' for cell in next(sheet.iter_rows(values_only=True))]
        recruiters = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            recruiters.append({
            'first_name': self._get_first_name(data),
            'email': self._get_email(data),
            'company_name': self._get_company_name(data)
            })
        return recruiters

    def _get_first_name(self, data):
        # Normalize keys: lowercase, remove spaces
        normalized = {k.lower().replace(" ", ""): v for k, v in data.items() if k and v}

        # Check common keys for first name
        for key in ['firstname', 'first_name']:
            if key in normalized:
                return str(normalized[key]).strip()

        # Try 'name' key for full name, return first part
        if 'name' in normalized and isinstance(normalized['name'], str):
            return normalized['name'].strip().split()[0]

        # Fallback: look for any key containing 'name'
        for key, value in normalized.items():
            if 'name' in key and isinstance(value, str):
                return value.strip().split()[0]

        return ''

    def _get_email(self, data):
        keys = {k.lower(): v for k, v in data.items()}
        return keys.get('email') or ''

    def _get_company_name(self, data):
        normalized = {k.lower().replace(" ", "").strip(): v for k, v in data.items() if k and v}
        for key in ['companyname', 'company', 'organization']:
            if key in normalized:
                return str(normalized[key]).strip()
        return ''
